#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
通用分镜信息表生成脚本
用于AI动漫制作工作流

使用方式：
1. 命令行调用：python generate_storyboard_table.py --data '<JSON数据>' --output '分镜信息表.xlsx'
2. 作为模块导入：from generate_storyboard_table import create_storyboard_table

JSON数据格式：
[
    {
        "episode": 1,               # 第几集
        "scene": 1,                 # 第几场
        "shot": 1,                  # 第几个镜头
        "duration": "3s",           # 时长(s)
        "shot_size": "中景",        # 景别：远景/全景/中景/近景/特写/大特写
        "camera": "固定",           # 摄法：固定/横摇/俯仰摇/推/拉/变焦/手持/跟随
        "content": "画面内容描述",   # 画面内容(≥50字，融合8大要素)
        "dialogue": "台词/音效",    # 台词/音效
        "characters": "角色A，角色B", # 入镜角色
        "scene_label": "场景名称",  # 场景标识
        "first_frame": "首帧提示词",
        "last_frame": "尾帧提示词",
        "video_prompt": "视频提示词"
    },
    ...
]
"""

import argparse
import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def create_storyboard_table(storyboards: list, output_path: str = "分镜信息表.xlsx") -> str:
    """
    创建分镜信息表Excel文件（完整13列版本）
    
    Args:
        storyboards: 分镜数据列表，每个元素包含 episode, scene, shot, duration,
                     shot_size, camera, content, dialogue, characters, scene_label,
                     first_frame, last_frame, video_prompt 字段
        output_path: 输出文件路径，默认为 "分镜信息表.xlsx"
    
    Returns:
        str: 生成的文件路径
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "分镜信息表"
    
    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    center_alignment = Alignment(horizontal='center', vertical='center')
    center_top_alignment = Alignment(horizontal='center', vertical='top')
    
    # 完整13列表头
    headers = [
        "第几集", "第几场", "第几个镜头", "时长(s)", "景别", "摄法",
        "画面内容", "台词/音效", "入镜角色", "场景标识",
        "首帧提示词", "尾帧提示词", "视频提示词"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_alignment
    
    # 填充数据
    for row, shot in enumerate(storyboards, 2):
        # 第几集
        ws.cell(row=row, column=1, value=shot.get("episode", 1)).border = thin_border
        ws.cell(row=row, column=1).alignment = center_top_alignment
        
        # 第几场
        ws.cell(row=row, column=2, value=shot.get("scene", 1)).border = thin_border
        ws.cell(row=row, column=2).alignment = center_top_alignment
        
        # 第几个镜头
        ws.cell(row=row, column=3, value=shot.get("shot", 1)).border = thin_border
        ws.cell(row=row, column=3).alignment = center_top_alignment
        
        # 时长(s)
        ws.cell(row=row, column=4, value=shot.get("duration", "")).border = thin_border
        ws.cell(row=row, column=4).alignment = center_top_alignment
        
        # 景别
        ws.cell(row=row, column=5, value=shot.get("shot_size", "")).border = thin_border
        ws.cell(row=row, column=5).alignment = center_top_alignment
        
        # 摄法
        ws.cell(row=row, column=6, value=shot.get("camera", "")).border = thin_border
        ws.cell(row=row, column=6).alignment = center_top_alignment
        
        # 画面内容
        ws.cell(row=row, column=7, value=shot.get("content", "")).border = thin_border
        ws.cell(row=row, column=7).alignment = wrap_alignment
        
        # 台词/音效
        ws.cell(row=row, column=8, value=shot.get("dialogue", "")).border = thin_border
        ws.cell(row=row, column=8).alignment = wrap_alignment
        
        # 入镜角色
        ws.cell(row=row, column=9, value=shot.get("characters", "")).border = thin_border
        ws.cell(row=row, column=9).alignment = center_top_alignment
        
        # 场景标识
        ws.cell(row=row, column=10, value=shot.get("scene_label", "")).border = thin_border
        ws.cell(row=row, column=10).alignment = center_top_alignment
        
        # 首帧提示词
        ws.cell(row=row, column=11, value=shot.get("first_frame", "")).border = thin_border
        ws.cell(row=row, column=11).alignment = wrap_alignment
        
        # 尾帧提示词
        ws.cell(row=row, column=12, value=shot.get("last_frame", "")).border = thin_border
        ws.cell(row=row, column=12).alignment = wrap_alignment
        
        # 视频提示词
        ws.cell(row=row, column=13, value=shot.get("video_prompt", "")).border = thin_border
        ws.cell(row=row, column=13).alignment = wrap_alignment
    
    # 设置列宽
    ws.column_dimensions['A'].width = 8    # 第几集
    ws.column_dimensions['B'].width = 8    # 第几场
    ws.column_dimensions['C'].width = 10   # 第几个镜头
    ws.column_dimensions['D'].width = 8    # 时长
    ws.column_dimensions['E'].width = 10   # 景别
    ws.column_dimensions['F'].width = 10   # 摄法
    ws.column_dimensions['G'].width = 55   # 画面内容
    ws.column_dimensions['H'].width = 30   # 台词/音效
    ws.column_dimensions['I'].width = 15   # 入镜角色
    ws.column_dimensions['J'].width = 15   # 场景标识
    ws.column_dimensions['K'].width = 45   # 首帧提示词
    ws.column_dimensions['L'].width = 45   # 尾帧提示词
    ws.column_dimensions['M'].width = 40   # 视频提示词
    
    # 设置行高
    for row in range(2, len(storyboards) + 2):
        ws.row_dimensions[row].height = 150
    
    wb.save(output_path)
    return output_path


def main():
    parser = argparse.ArgumentParser(
        description='生成分镜信息表Excel文件（完整13列版本）',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
示例:
  python generate_storyboard_table.py --data '[{"episode":1,"scene":1,"shot":1,"duration":"3s","shot_size":"中景","camera":"固定","content":"画面内容","dialogue":"台词","characters":"角色A","scene_label":"场景名","first_frame":"首帧","last_frame":"尾帧","video_prompt":"视频"}]' --output 分镜信息表.xlsx
  python generate_storyboard_table.py --file storyboards.json --output 分镜信息表.xlsx
        '''
    )
    
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument('--data', type=str, help='JSON格式的分镜数据字符串')
    group.add_argument('--file', type=str, help='包含分镜数据的JSON文件路径')
    
    parser.add_argument('--output', type=str, default='分镜信息表.xlsx', help='输出文件路径（默认：分镜信息表.xlsx）')
    
    args = parser.parse_args()
    
    # 解析数据
    try:
        if args.data:
            storyboards = json.loads(args.data)
        else:
            with open(args.file, 'r', encoding='utf-8') as f:
                storyboards = json.load(f)
    except json.JSONDecodeError as e:
        print(f"错误：JSON解析失败 - {e}", file=sys.stderr)
        sys.exit(1)
    except FileNotFoundError:
        print(f"错误：文件不存在 - {args.file}", file=sys.stderr)
        sys.exit(1)
    
    # 验证数据格式
    if not isinstance(storyboards, list):
        print("错误：数据必须是数组格式", file=sys.stderr)
        sys.exit(1)
    
    # 生成表格
    output_file = create_storyboard_table(storyboards, args.output)
    print(f"✅ 分镜信息表已生成：{output_file}")
    
    # 统计信息
    episodes = set(s.get("episode", 1) for s in storyboards)
    print(f"   共 {len(episodes)} 集，{len(storyboards)} 个分镜")


if __name__ == "__main__":
    main()
