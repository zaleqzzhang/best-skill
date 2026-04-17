#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
通用人物信息表生成脚本
用于AI动漫制作工作流

使用方式：
1. 命令行调用：python generate_character_table.py --data '<JSON数据>' --output '人物信息表.xlsx'
2. 作为模块导入：from generate_character_table import create_character_table

JSON数据格式：
[
    {
        "id": "C01",
        "name": "角色名",
        "info": "人物信息提取内容",
        "bio": "人物小传内容",
        "prompt": "三视图提示词"
    },
    ...
]
"""

import argparse
import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def create_character_table(characters: list, output_path: str = "人物信息表.xlsx") -> str:
    """
    创建人物信息表Excel文件
    
    Args:
        characters: 人物数据列表，每个元素包含 id, name, info, bio, prompt 字段
        output_path: 输出文件路径，默认为 "人物信息表.xlsx"
    
    Returns:
        str: 生成的文件路径
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "人物信息表"
    
    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # 表头
    headers = ["人物编号", "人物姓名", "人物信息提取", "人物小传", "人物三视图提示词"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_alignment
    
    # 填充数据
    for row, char in enumerate(characters, 2):
        # 人物编号
        ws.cell(row=row, column=1, value=char.get("id", f"C{row-1:02d}")).border = thin_border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='top')
        
        # 人物姓名
        ws.cell(row=row, column=2, value=char.get("name", "")).border = thin_border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center', vertical='top')
        
        # 人物信息提取
        ws.cell(row=row, column=3, value=char.get("info", "")).border = thin_border
        ws.cell(row=row, column=3).alignment = wrap_alignment
        
        # 人物小传
        ws.cell(row=row, column=4, value=char.get("bio", "")).border = thin_border
        ws.cell(row=row, column=4).alignment = wrap_alignment
        
        # 三视图提示词
        ws.cell(row=row, column=5, value=char.get("prompt", "")).border = thin_border
        ws.cell(row=row, column=5).alignment = wrap_alignment
    
    # 设置列宽
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 60
    
    # 设置行高（根据内容自适应）
    for row in range(2, len(characters) + 2):
        ws.row_dimensions[row].height = 200
    
    wb.save(output_path)
    return output_path


def main():
    parser = argparse.ArgumentParser(
        description='生成人物信息表Excel文件',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
示例:
  python generate_character_table.py --data '[{"id":"C01","name":"主角","info":"主角信息","bio":"主角小传","prompt":"提示词"}]' --output 人物信息表.xlsx
  python generate_character_table.py --file characters.json --output 人物信息表.xlsx
        '''
    )
    
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument('--data', type=str, help='JSON格式的人物数据字符串')
    group.add_argument('--file', type=str, help='包含人物数据的JSON文件路径')
    
    parser.add_argument('--output', type=str, default='人物信息表.xlsx', help='输出文件路径（默认：人物信息表.xlsx）')
    
    args = parser.parse_args()
    
    # 解析数据
    try:
        if args.data:
            characters = json.loads(args.data)
        else:
            with open(args.file, 'r', encoding='utf-8') as f:
                characters = json.load(f)
    except json.JSONDecodeError as e:
        print(f"错误：JSON解析失败 - {e}", file=sys.stderr)
        sys.exit(1)
    except FileNotFoundError:
        print(f"错误：文件不存在 - {args.file}", file=sys.stderr)
        sys.exit(1)
    
    # 验证数据格式
    if not isinstance(characters, list):
        print("错误：数据必须是数组格式", file=sys.stderr)
        sys.exit(1)
    
    # 生成表格
    output_file = create_character_table(characters, args.output)
    print(f"✅ 人物信息表已生成：{output_file}")
    print(f"   共 {len(characters)} 个角色")


if __name__ == "__main__":
    main()
