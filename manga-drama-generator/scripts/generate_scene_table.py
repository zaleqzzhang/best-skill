#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
通用场景信息表生成脚本
用于AI动漫制作工作流

使用方式：
1. 命令行调用：python generate_scene_table.py --data '<JSON数据>' --output '场景信息表.xlsx'
2. 作为模块导入：from generate_scene_table import create_scene_table

JSON数据格式：
[
    {
        "id": "S01",
        "name": "场景名称",
        "prompt": "场景提示词"
    },
    ...
]
"""

import argparse
import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def create_scene_table(scenes: list, output_path: str = "场景信息表.xlsx") -> str:
    """
    创建场景信息表Excel文件
    
    Args:
        scenes: 场景数据列表，每个元素包含 id, name, prompt 字段
        output_path: 输出文件路径，默认为 "场景信息表.xlsx"
    
    Returns:
        str: 生成的文件路径
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "场景信息表"
    
    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # 表头
    headers = ["场景编号", "场景名字", "场景提示词"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_alignment
    
    # 填充数据
    for row, scene in enumerate(scenes, 2):
        # 场景编号
        ws.cell(row=row, column=1, value=scene.get("id", f"S{row-1:02d}")).border = thin_border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='top')
        
        # 场景名字
        ws.cell(row=row, column=2, value=scene.get("name", "")).border = thin_border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center', vertical='top')
        
        # 场景提示词
        ws.cell(row=row, column=3, value=scene.get("prompt", "")).border = thin_border
        ws.cell(row=row, column=3).alignment = wrap_alignment
    
    # 设置列宽
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 80
    
    # 设置行高
    for row in range(2, len(scenes) + 2):
        ws.row_dimensions[row].height = 120
    
    wb.save(output_path)
    return output_path


def main():
    parser = argparse.ArgumentParser(
        description='生成场景信息表Excel文件',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
示例:
  python generate_scene_table.py --data '[{"id":"S01","name":"客厅","prompt":"场景提示词内容"}]' --output 场景信息表.xlsx
  python generate_scene_table.py --file scenes.json --output 场景信息表.xlsx
        '''
    )
    
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument('--data', type=str, help='JSON格式的场景数据字符串')
    group.add_argument('--file', type=str, help='包含场景数据的JSON文件路径')
    
    parser.add_argument('--output', type=str, default='场景信息表.xlsx', help='输出文件路径（默认：场景信息表.xlsx）')
    
    args = parser.parse_args()
    
    # 解析数据
    try:
        if args.data:
            scenes = json.loads(args.data)
        else:
            with open(args.file, 'r', encoding='utf-8') as f:
                scenes = json.load(f)
    except json.JSONDecodeError as e:
        print(f"错误：JSON解析失败 - {e}", file=sys.stderr)
        sys.exit(1)
    except FileNotFoundError:
        print(f"错误：文件不存在 - {args.file}", file=sys.stderr)
        sys.exit(1)
    
    # 验证数据格式
    if not isinstance(scenes, list):
        print("错误：数据必须是数组格式", file=sys.stderr)
        sys.exit(1)
    
    # 生成表格
    output_file = create_scene_table(scenes, args.output)
    print(f"✅ 场景信息表已生成：{output_file}")
    print(f"   共 {len(scenes)} 个场景")


if __name__ == "__main__":
    main()
